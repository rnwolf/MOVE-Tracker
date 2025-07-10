#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "pandas>=2.0",
#     "numpy>=1.20",
#     "matplotlib>=3.5",
#     "openpyxl>=3.0",
#     "rich>=12.0",
#     "typer>=0.9.0",
#     "python-dateutil>=2.8.2",
#     "scikit-learn>=1.7.0",
# ]
# ///

import logging
import os
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from enum import Enum
from typing import Optional
from sklearn.linear_model import LinearRegression

import typer
from rich.console import Console
from rich.prompt import Confirm
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from dateutil.relativedelta import relativedelta
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt

# ==============================================================================
# Data Structures
# ==============================================================================


class BufferSignal(Enum):
    GREEN = "Green"
    YELLOW = "Yellow"
    RED = "Red"
    BEYOND_RED = "Beyond Red"


@dataclass
class MOVEConfiguration:
    planned_start_date: date
    planned_delivery_date: date
    buffer_green_date: date
    buffer_yellow_date: date
    buffer_red_date: date
    buffer_beyond_red_date: date
    fever_green_yellow_left_y: float
    fever_green_yellow_right_y: float
    fever_yellow_red_left_y: float
    fever_yellow_red_right_y: float
    historic_50th_percentile_flow_time_override: Optional[float] = None


# ==============================================================================
# CLI Application
# ==============================================================================

app = typer.Typer()
console = Console()

# ==============================================================================
# Core Functions
# ==============================================================================


def _setup_logging(log_level: str):
    """Set up logging."""
    logging.basicConfig(
        level=log_level.upper(),
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _create_excel_template(excel_path: str, overwrite: bool):
    """Creates a new Excel file with the required sheets and headers."""
    if os.path.exists(excel_path):
        if not overwrite:
            confirm = Confirm.ask(
                f"[bold yellow]File '{excel_path}' already exists. Overwrite?[/bold yellow]",
                default=False,
            )
            if not confirm:
                console.print("[bold red]Operation cancelled by user.[/bold red]")
                raise typer.Exit(code=0)
        else:
            console.print(
                f"[bold yellow]Overwriting existing file: {excel_path}[/bold yellow]"
            )
    try:
        console.print(
            f"Creating Excel template at: [bold cyan]{excel_path}[/bold cyan]"
        )

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        # Define planned dates
        today = date.today()
        # Calculate the first day of the next month
        if today.month == 12:
            planned_start_date = date(today.year + 1, 1, 1)
        else:
            planned_start_date = date(today.year, today.month + 1, 1)

        # Calculate the end of the month, three months from planned_start_date
        from datetime import timedelta

        planned_delivery_date = (
            planned_start_date + relativedelta(months=+1) - timedelta(days=1)
        )

        # Instructions Sheet
        ws_instructions: Worksheet = wb.create_sheet("Instructions")
        ws_instructions["A1"] = "Instructions for using the MOVE Tracker"
        ws_instructions["A3"] = (
            "1. Populate 'Historic_Work_Items' with data from past projects to calculate baseline flow time."
        )
        ws_instructions["A4"] = (
            "2. Populate 'Current_Work_Items' with the work items for this project."
        )
        ws_instructions["A5"] = (
            "3. Populate 'MOVE_Configuration' with the project's parameters."
        )
        ws_instructions["A6"] = (
            "4. Do not delete items from 'Current_Work_Items'. Instead, mark them with a 'Date_Withdrawn'."
        )
        ws_instructions["A7"] = "5. Run the script to generate reports."

        # Historic_Work_Items Sheet
        ws_historic: Worksheet = wb.create_sheet("Historic_Work_Items")
        historic_headers = [
            "Historical_WI_ID",
            "Description",
            "Actual_Start_Date",
            "Actual_Completion_Date",
            "Flow_Time_Days",
        ]
        ws_historic.append(historic_headers)

        # Add 20 historic sample data items
        from datetime import timedelta
        import random

        today = datetime.now()
        for i in range(1, 21):
            wi_id = f"HIST-WI-{i:03d}"
            description = f"Sample Historic Work Item {i}"
            flow_time_days = random.randint(2, 10)
            actual_completion_date = today - timedelta(days=random.randint(1, 365))
            actual_start_date = actual_completion_date - timedelta(days=flow_time_days)

            # Excel formula for Flow_Time_Days
            # Assuming Actual_Start_Date is in column C and Actual_Completion_Date is in column D
            # The row number will be i + 1 (for header) + 1 (for 0-indexed loop) = i + 2
            flow_time_formula = f"=INT(D{i+1})-INT(C{i+1})+1"

            ws_historic.append(
                [
                    wi_id,
                    description,
                    actual_start_date.date(),
                    actual_completion_date.date(),
                    flow_time_formula,
                ]
            )

        # Current_Work_Items Sheet
        ws_current: Worksheet = wb.create_sheet("Current_Work_Items")
        current_headers = [
            "Work_Item_ID",
            "Description",
            "Commitment_Date",
            "Status",
            "Actual_Start_Date",
            "Actual_Completion_Date",
            "Date_Withdrawn",
        ]
        ws_current.append(current_headers)

        # Add 8 sample current work items
        last_completion_date = planned_start_date  # Initialize with planned_start_date

        for i in range(1, 9):  # 8 items
            wi_id = f"WI-{i:03d}"
            description = f"Sample Current Work Item {i}"
            status = "Completed"  # All completed for now
            if i <= 6:
                commitment_date = planned_start_date
            else:
                commitment_date = planned_start_date + relativedelta(months=+1)

            actual_start_date = last_completion_date

            # Random flow time between 2 and 10 days
            flow_time = random.randint(2, 10)
            actual_completion_date = actual_start_date + timedelta(days=flow_time)

            date_withdrawn = ""  # Leave empty

            ws_current.append(
                [
                    wi_id,
                    description,
                    commitment_date,
                    status,
                    actual_start_date,
                    actual_completion_date,
                    date_withdrawn,
                ]
            )

            last_completion_date = actual_completion_date  # Update for next iteration

        # Convert date strings to actual dates
        # ws_current["C:C"].number_format = "YYYY-MM-DD"
        # ws_current["E:E"].number_format = "YYYY-MM-DD"
        # ws_current["F:F"].number_format = "YYYY-MM-DD"
        # ws_current["G:G"].number_format = "YYYY-MM-DD"

        # MOVE_Configuration Sheet
        ws_config: Worksheet = wb.create_sheet("MOVE_Configuration")
        config_headers = ["Parameter", "Value"]
        ws_config.append(config_headers)

        config_data = [
            ("Planned_Start_Date", planned_start_date),
            ("Planned_Delivery_Date", planned_delivery_date),
            (
                "Historic_50th_Percentile_Flow_Time",
                "=_xlfn.PERCENTILE.INC('Historic_Work_Items'!E:E, 0.5)",
            ),
            ("Historic_50th_Percentile_Flow_Time_Override", ""),
            (
                "Initial_Scope",
                "=SUMPRODUCT( --(Current_Work_Items!C$2:INDEX(Current_Work_Items!C:C, MATCH(1E+100, Current_Work_Items!C:C)) <= $B$2), --((Current_Work_Items!G$2:INDEX(Current_Work_Items!G:G, MATCH(1E+100, Current_Work_Items!C:C)) > $B$2) + ISBLANK(Current_Work_Items!G$2:INDEX(Current_Work_Items!G:G, MATCH(1E+100,Current_Work_Items!C:C)))))",
            ),
            ("Initial_Ideal_Completion_Flow_Time", "=B6*B4"),
            ("Buffer_Green_Date", "=B2+B7"),
            ("Buffer_Yellow_Date", "=B8+(0.2*B7)"),
            ("Buffer_Red_Date", "=B9+(0.2*B7)"),
            ("Buffer_Beyond_Red_Date", "=B10+(0.2*B7)"),
            ("Fever_Green_Yellow_Left_Y", 0.2),
            ("Fever_Green_Yellow_Right_Y", 0.5),
            ("Fever_Yellow_Red_Left_Y", 0.5),
            ("Fever_Yellow_Red_Right_Y", 0.8),
        ]
        for row in config_data:
            ws_config.append(row)

        # Convert date strings to actual dates
        ws_config["B2"].number_format = "YYYY-MM-DD"
        ws_config["B3"].number_format = "YYYY-MM-DD"
        # Apply date formatting to buffer dates
        ws_config["B8"].number_format = "YYYY-MM-DD"
        ws_config["B9"].number_format = "YYYY-MM-DD"
        ws_config["B10"].number_format = "YYYY-MM-DD"
        ws_config["B11"].number_format = "YYYY-MM-DD"

        # Progress_Log Sheet
        ws_progress: Worksheet = wb.create_sheet("Progress_Log")
        progress_headers = [
            "Snapshot_Date",
            "Scope_At_Snapshot",
            "Actual_Work_Completed",
            "Elapsed_Time_Days",
            "Actual_Operational_Throughput",
            "Current_50th_Percentile_Flow_Time",
            "Forecasted_Delivery_Date",
            "Buffer_Consumption_Percentage",
            "Work_Done_Percentage",
            "Current_Buffer_Signal",
        ]
        ws_progress.append(progress_headers)

        # Chart Sheets
        wb.create_sheet("Work_Execution_Chart")
        wb.create_sheet("Fever_Chart")

        wb.save(excel_path)
        console.print(
            f"[bold green]Successfully created template: {excel_path}[/bold green]"
        )

    except Exception as e:
        console.print(
            f"[bold red]An unexpected error occurred during template creation: {e}[/bold red]"
        )
        logging.error(f"Template creation failed: {e}", exc_info=True)
        raise typer.Exit(code=1)


def _read_excel_data(
    excel_path: str,
) -> tuple[MOVEConfiguration, pd.DataFrame, pd.DataFrame]:
    """Reads and validates data from the Excel workbook."""
    logging.info(f"Reading Excel data from: {excel_path}")
    try:
        with pd.ExcelFile(excel_path) as xls:
            if "MOVE_Configuration" not in xls.sheet_names:
                console.print(
                    "[bold red]Error: 'MOVE_Configuration' sheet not found in the Excel file.[/bold red]"
                )
                raise typer.Exit(code=1)
            df_config = pd.read_excel(
                xls, "MOVE_Configuration", engine="openpyxl"
            ).set_index("Parameter")["Value"]
            print(df_config)
            # --- Parse and validate MOVEConfiguration ---
            config_dict = {}
            required_params = [
                "Planned_Start_Date",
                "Planned_Delivery_Date",
                "Buffer_Green_Date",
                "Buffer_Yellow_Date",
                "Buffer_Red_Date",
                "Buffer_Beyond_Red_Date",
                "Fever_Green_Yellow_Left_Y",
                "Fever_Green_Yellow_Right_Y",
                "Fever_Yellow_Red_Left_Y",
                "Fever_Yellow_Red_Right_Y",
            ]
            for param in required_params:
                if param not in df_config.index or pd.isna(df_config.get(param)):
                    console.print(
                        f"[bold red]Error: Required configuration parameter '{param}' is missing or empty in 'MOVE_Configuration' sheet.[/bold red]"
                    )
                    raise typer.Exit(code=1)

            date_params = [
                "Planned_Start_Date",
                "Planned_Delivery_Date",
                "Buffer_Green_Date",
                "Buffer_Yellow_Date",
                "Buffer_Red_Date",
                "Buffer_Beyond_Red_Date",
            ]
            for param in date_params:
                config_dict[param.lower()] = pd.to_datetime(df_config[param]).date()

            float_params = [
                "Fever_Green_Yellow_Left_Y",
                "Fever_Green_Yellow_Right_Y",
                "Fever_Yellow_Red_Left_Y",
                "Fever_Yellow_Red_Right_Y",
            ]
            for param in float_params:
                config_dict[param.lower()] = float(df_config[param])

            # Optional override
            override_val = df_config.get("Historic_50th_Percentile_Flow_Time_Override")
            config_dict["historic_50th_percentile_flow_time_override"] = (
                float(override_val) if pd.notna(override_val) else None
            )

            move_config = MOVEConfiguration(**config_dict)
            logging.debug(f"MOVE Configuration loaded: {move_config}")

            # --- Read Work Item Sheets ---
            sheet_to_df = {}
            required_sheets = {
                "Historic_Work_Items": [
                    "Historical_WI_ID",
                    "Actual_Start_Date",
                    "Actual_Completion_Date",
                    "Flow_Time_Days",
                ],
                "Current_Work_Items": [
                    "Work_Item_ID",
                    "Status",
                    "Actual_Start_Date",
                    "Actual_Completion_Date",
                ],
            }

            for sheet, required_cols in required_sheets.items():
                if sheet not in xls.sheet_names:
                    console.print(
                        f"[bold red]Error: '{sheet}' sheet not found in the Excel file.[/bold red]"
                    )
                    raise typer.Exit(code=1)
                df = pd.read_excel(xls, sheet, engine="openpyxl")
                if not all(col in df.columns for col in required_cols):
                    console.print(
                        f"[bold red]Error: Missing one or more required columns in '{sheet}'. Required: {required_cols}[/bold red]"
                    )
                    raise typer.Exit(code=1)
                sheet_to_df[sheet] = df

            df_historic = sheet_to_df["Historic_Work_Items"]
            df_current = sheet_to_df["Current_Work_Items"]

            # Convert date columns
            for col in ["Actual_Start_Date", "Actual_Completion_Date"]:
                if col in df_historic.columns:
                    df_historic[col] = pd.to_datetime(df_historic[col], errors="coerce")
            for col in [
                "Actual_Start_Date",
                "Actual_Completion_Date",
                "Date_Withdrawn",
            ]:
                if col in df_current.columns:
                    df_current[col] = pd.to_datetime(df_current[col], errors="coerce")

            console.print(
                "[bold green]Successfully read and validated all input data.[/bold green]"
            )
            return move_config, df_historic, df_current

    except FileNotFoundError:
        console.print(
            f"[bold red]Error: The file at '{excel_path}' was not found.[/bold red]"
        )
        raise typer.Exit(code=1)
    except Exception as e:
        console.print(
            f"[bold red]An unexpected error occurred while reading the Excel file: {e}[/bold red]"
        )
        logging.error(f"Excel reading failed: {e}", exc_info=True)
        raise typer.Exit(code=1)


def _generate_full_progress_log(
    move_config: MOVEConfiguration,
    df_current: pd.DataFrame,
    historic_50th_percentile_flow_time: float,
    snapshot_date_str: str,
) -> pd.DataFrame:
    """Generates the historical Progress_Log from start date to snapshot date."""
    try:
        # Work with Timestamps for consistency
        snapshot_ts = pd.to_datetime(snapshot_date_str)
        planned_start_ts = pd.to_datetime(move_config.planned_start_date)

        if snapshot_ts < planned_start_ts:
            console.print(
                "[bold red]Error: Snapshot date cannot be before the planned start date.[/bold red]"
            )
            raise typer.Exit(code=1)
    except ValueError:
        console.print(
            f"[bold red]Error: Invalid snapshot date format '{snapshot_date_str}'. Please use YYYY-MM-DD.[/bold red]"
        )
        raise typer.Exit(code=1)

    logging.info(
        f"Generating progress log from {planned_start_ts.date()} to {snapshot_ts.date()}"
    )

    # Identify all relevant "event dates" from Current_Work_Items
    event_dates = set()
    event_dates.add(planned_start_ts.date())  # Add planned start date
    event_dates.add(snapshot_ts.date())  # Add snapshot date

    # Add all unique dates from Current_Work_Items that are relevant
    for col in [
        "Commitment_Date",
        "Actual_Start_Date",
        "Actual_Completion_Date",
        "Date_Withdrawn",
    ]:
        if col in df_current.columns:
            # Filter out NaT (Not a Time) values before adding to set
            valid_dates = (
                df_current[col].dropna().apply(lambda x: pd.to_datetime(x).date())
            )
            event_dates.update(valid_dates)

    # Filter dates to be within the planned_start_ts and snapshot_ts range
    # and convert to Timestamps for consistency with existing logic
    all_dates = sorted(
        [
            pd.Timestamp(d)
            for d in event_dates
            if planned_start_ts.date() <= d <= snapshot_ts.date()
        ]
    )

    # Convert config dates to Timestamps for comparison
    planned_delivery_ts = pd.to_datetime(move_config.planned_delivery_date)
    buffer_red_ts = pd.to_datetime(move_config.buffer_red_date)

    progress_log_entries = []

    for current_ts in all_dates:  # current_ts is a Timestamp
        log_entry = {}

        # Calculate Scope_At_Snapshot (compare datetime64[ns] with Timestamp)
        scope_mask = (df_current["Commitment_Date"] <= current_ts) & (
            (df_current["Date_Withdrawn"].isnull())
            | (df_current["Date_Withdrawn"] > current_ts)
        )
        scope_at_snapshot = scope_mask.sum()

        # Calculate Actual_Work_Completed
        completed_mask = (df_current["Status"] == "Completed") & (
            df_current["Actual_Completion_Date"] <= current_ts
        )
        actual_work_completed = completed_mask.sum()

        # Other calculations
        elapsed_time_days = (current_ts - planned_start_ts).days + 1
        actual_operational_throughput = (
            actual_work_completed / elapsed_time_days if elapsed_time_days > 0 else 0
        )

        # Current 50th Percentile Flow Time
        df_completed_current = df_current[completed_mask]
        if len(df_completed_current) >= 2:
            flow_times = (
                df_completed_current["Actual_Completion_Date"]
                - df_completed_current["Actual_Start_Date"]
            ).dt.days + 1
            current_50th_percentile_flow_time = np.percentile(flow_times.dropna(), 50)
        else:
            current_50th_percentile_flow_time = historic_50th_percentile_flow_time

        # Collect historical data points for regression up to current_ts
        # Filter progress_log_entries to only include entries up to current_ts
        historical_progress_data = [
            entry
            for entry in progress_log_entries
            if pd.Timestamp(entry["Snapshot_Date"]) <= current_ts
        ]

        # Add the current_ts data point if it's not already there (it will be the last one)
        # This ensures we have the most up-to-date actual_work_completed for the current_ts
        # We need to ensure that the current_ts data point is included in the regression
        # even if it's the first entry or if there are no prior entries.
        # For the regression, we need (elapsed_time_days, actual_work_completed)

        # Create a list of (elapsed_time_days, actual_work_completed) for regression
        regression_data = []
        # Add previous entries
        for entry in historical_progress_data:
            regression_data.append(
                (entry["Elapsed_Time_Days"], entry["Actual_Work_Completed"])
            )
        # Add current entry
        regression_data.append((elapsed_time_days, actual_work_completed))

        # Ensure unique elapsed_time_days for regression
        # If multiple entries have the same elapsed_time_days, take the last one (most recent)
        unique_regression_data = {}
        for days, completed in regression_data:
            unique_regression_data[days] = completed

        sorted_regression_data = sorted(unique_regression_data.items())

        X = np.array([item[0] for item in sorted_regression_data]).reshape(-1, 1)
        y = np.array([item[1] for item in sorted_regression_data])

        # Forecasted Delivery Date using Linear Regression
        if len(X) >= 2:  # Need at least 2 data points for linear regression
            model = LinearRegression()
            model.fit(X, y)
            slope = model.coef_[0]
            intercept = model.intercept_

            if slope > 0:
                # Solve: slope * x + intercept = scope_at_snapshot
                # x = (scope_at_snapshot - intercept) / slope
                days_to_reach_scope = (scope_at_snapshot - intercept) / slope
                forecasted_delivery_date = planned_start_ts + pd.to_timedelta(
                    days_to_reach_scope, unit="D"
                )
            else:
                forecasted_delivery_date = (
                    pd.NaT
                )  # Cannot forecast with non-positive slope
        else:
            forecasted_delivery_date = pd.NaT  # Not enough data for regression

        # Buffer Consumption
        if pd.notna(forecasted_delivery_date):
            # Compare Timestamps
            # Convert config dates to Timestamps for comparison
            buffer_green_ts = pd.to_datetime(move_config.buffer_green_date)
            buffer_beyond_red_ts = pd.to_datetime(move_config.buffer_beyond_red_date)

            buffer_delta = (buffer_beyond_red_ts - buffer_green_ts).days
            forecast_delta = (forecasted_delivery_date - buffer_green_ts).days
            buffer_consumption_percentage = (
                forecast_delta / buffer_delta if buffer_delta > 0 else 0
            )
        else:
            buffer_consumption_percentage = 0
        buffer_consumption_percentage = max(
            0, buffer_consumption_percentage
        )  # Cap at 0 if ahead

        # Work Done Percentage
        work_done_percentage = (
            actual_work_completed / scope_at_snapshot if scope_at_snapshot > 0 else 0
        )

        # Current Buffer Signal - Here we need to compare dates, not timestamps
        if pd.notna(forecasted_delivery_date):
            fdd = (
                forecasted_delivery_date.date()
            )  # Convert to date for comparison with config dates
            if fdd <= move_config.buffer_green_date:
                current_buffer_signal = BufferSignal.GREEN.value
            elif fdd <= move_config.buffer_yellow_date:
                current_buffer_signal = BufferSignal.YELLOW.value
            elif fdd <= move_config.buffer_red_date:
                current_buffer_signal = BufferSignal.RED.value
            else:
                current_buffer_signal = BufferSignal.BEYOND_RED.value
        else:
            current_buffer_signal = BufferSignal.GREEN.value  # Default if no forecast

        log_entry["Snapshot_Date"] = current_ts  # Store the Timestamp
        log_entry["Scope_At_Snapshot"] = scope_at_snapshot
        log_entry["Actual_Work_Completed"] = actual_work_completed
        log_entry["Elapsed_Time_Days"] = elapsed_time_days
        log_entry["Actual_Operational_Throughput"] = actual_operational_throughput
        log_entry["Current_50th_Percentile_Flow_Time"] = (
            current_50th_percentile_flow_time
        )
        log_entry["Forecasted_Delivery_Date"] = forecasted_delivery_date
        log_entry["Buffer_Consumption_Percentage"] = buffer_consumption_percentage
        log_entry["Work_Done_Percentage"] = work_done_percentage
        log_entry["Current_Buffer_Signal"] = current_buffer_signal

        progress_log_entries.append(log_entry)

    if not progress_log_entries:
        console.print(
            "[bold yellow]Warning: No progress log entries were generated.[/bold yellow]"
        )
        return pd.DataFrame()

    df_progress_log = pd.DataFrame(progress_log_entries)
    # The 'Snapshot_Date' column is already datetime64[ns], so this conversion is not strictly needed but harmless.
    df_progress_log["Snapshot_Date"] = pd.to_datetime(df_progress_log["Snapshot_Date"])

    console.print(
        "[bold green]Successfully generated the full progress log.[/bold green]"
    )
    logging.debug(f"Generated Progress Log Head:\n{df_progress_log.head()}")
    return df_progress_log


def _update_progress_log_sheet(excel_path: str, df_progress_log: pd.DataFrame):
    """Updates the Progress_Log sheet in the Excel file."""
    if df_progress_log.empty:
        logging.warning("Progress log is empty. Skipping sheet update.")
        return

    logging.info(f"Updating 'Progress_Log' sheet in {excel_path}")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        book = load_workbook(excel_path)
        if "Progress_Log" in book.sheetnames:
            sheet = book["Progress_Log"]
            # Clear existing data, keeping headers
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
        else:
            sheet = book.create_sheet("Progress_Log")
            # If sheet is new, headers will be written by dataframe_to_rows

        # Write the new data
        rows = dataframe_to_rows(df_progress_log, index=False, header=True)

        # If the sheet was cleared, we need to re-append headers
        # A simple approach is to just clear it all and write from scratch
        if "Progress_Log" in book.sheetnames:
            del book["Progress_Log"]
        sheet = book.create_sheet(
            "Progress_Log", -2
        )  # Place it before the chart sheets

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        book.save(excel_path)
        console.print(
            "[bold green]Successfully updated the 'Progress_Log' sheet.[/bold green]"
        )

    except Exception as e:
        console.print(
            f"[bold red]An error occurred while updating the Progress_Log sheet: {e}[/bold red]"
        )
        logging.error(f"Failed to update Excel sheet: {e}", exc_info=True)
        raise typer.Exit(code=1)


def _generate_charts(
    df_progress_log: pd.DataFrame,
    move_config: MOVEConfiguration,
    snapshot_date_str: str,
) -> Optional[tuple[str, str]]:
    """Generates Work Execution and Fever charts."""
    if df_progress_log.empty:
        logging.warning("Progress log is empty. Skipping chart generation.")
        return None

    logging.info("Generating charts...")
    snapshot_date = pd.to_datetime(snapshot_date_str).date()

    # --- Work Execution Chart ---
    fig_we, ax_we = plt.subplots(figsize=(12, 8))
    ax_we.set_title(f"Work Execution Signal Chart as at {snapshot_date_str}")

    # Plot data
    ax_we.plot(
        df_progress_log["Snapshot_Date"],
        df_progress_log["Actual_Work_Completed"],
        label="Actual Work Completed",
        marker="o",
        linestyle="-",
    )
    ax_we.plot(
        df_progress_log["Snapshot_Date"],
        df_progress_log["Scope_At_Snapshot"],
        label="Scope At Snapshot",
        linestyle="--",
    )

    # Trendline
    last_snapshot_data = df_progress_log.iloc[-1]
    x_numeric = (
        df_progress_log["Snapshot_Date"] - df_progress_log["Snapshot_Date"].min()
    ).dt.days
    if len(x_numeric) > 1:
        m, b = np.polyfit(x_numeric, df_progress_log["Actual_Work_Completed"], 1)
        trendline_x = pd.date_range(
            start=df_progress_log["Snapshot_Date"].min(),
            periods=len(df_progress_log) + 90,
        )
        trendline_y = (
            m * (trendline_x - df_progress_log["Snapshot_Date"].min()).days + b
        )
        ax_we.plot(trendline_x, trendline_y, "r--", label="Completion Trend")

    # Extended Scope Line
    last_scope = last_snapshot_data["Scope_At_Snapshot"]
    ax_we.axhline(
        y=last_scope,
        color="grey",
        linestyle="--",
        label=f"Current Scope ({last_scope})",
    )

    # Vertical Buffer Lines
    ax_we.axvline(
        x=move_config.planned_delivery_date,
        color="blue",
        linestyle="--",
        label=f"Planned Delivery: {move_config.planned_delivery_date.strftime('%Y-%m-%d')}",
    )
    # ax_we.axvline(
    #     x=move_config.buffer_green_date,
    #     color="green",
    #     linestyle="--",
    #     label="Green Buffer",
    # )
    # ax_we.axvline(
    #     x=move_config.buffer_yellow_date,
    #     color="yellow",
    #     linestyle="--",
    #     label="Yellow Buffer",
    # )

    ax_we.axvline(
        move_config.buffer_green_date,
        color="green",
        linestyle="--",
        label=f"Green Buffer Start: {move_config.buffer_green_date.strftime('%Y-%m-%d')}",
    )
    ax_we.axvline(
        move_config.buffer_yellow_date,
        color="orange",
        linestyle="--",
        label=f"Yellow Buffer Start: {move_config.buffer_yellow_date.strftime('%Y-%m-%d')}",
    )
    ax_we.axvline(
        move_config.buffer_red_date,
        color="red",
        linestyle="--",
        label=f"Red Buffer Start: {move_config.buffer_red_date.strftime('%Y-%m-%d')}",
    )
    ax_we.axvline(
        move_config.buffer_beyond_red_date,
        color="darkred",
        linestyle="--",
        label=f"Beyond Red Buffer Start: {move_config.buffer_beyond_red_date.strftime('%Y-%m-%d')}",
    )

    # Forecast Intersection
    forecasted_date = last_snapshot_data["Forecasted_Delivery_Date"]
    if pd.notna(forecasted_date):
        ax_we.plot(
            forecasted_date,
            last_scope,
            "X",
            color="purple",
            markersize=12,
            label=f'Forecast: {forecasted_date.strftime("%Y-%m-%d")}',
        )

    ax_we.set_xlabel("Date")
    ax_we.set_ylabel("Work Items")
    ax_we.legend()
    ax_we.grid(True)
    plt.tight_layout()
    work_execution_chart_path = f"{snapshot_date_str}_work_execution_chart.png"
    fig_we.savefig(work_execution_chart_path)
    plt.close(fig_we)
    console.print(
        f"[green]Work Execution Chart saved to {work_execution_chart_path}[/green]"
    )

    # --- Fever Chart ---
    fig_fever, ax_fever = plt.subplots(figsize=(10, 8))
    ax_fever.set_title(f"Fever Chart as at {snapshot_date_str}")

    # Background zones
    ax_fever.fill_between(
        [0, 1],
        [move_config.fever_green_yellow_left_y, move_config.fever_green_yellow_right_y],
        0,
        color="lightgreen",
        alpha=0.5,
        label="Green Zone",
    )
    ax_fever.fill_between(
        [0, 1],
        [move_config.fever_green_yellow_left_y, move_config.fever_green_yellow_right_y],
        [move_config.fever_yellow_red_left_y, move_config.fever_yellow_red_right_y],
        color="khaki",
        alpha=0.5,
        label="Yellow Zone",
    )
    ax_fever.fill_between(
        [0, 1],
        [move_config.fever_yellow_red_left_y, move_config.fever_yellow_red_right_y],
        2,
        color="lightcoral",
        alpha=0.5,
        label="Red Zone",
    )  # y2=2 to fill up

    # Plot data
    ax_fever.plot(
        df_progress_log["Work_Done_Percentage"],
        df_progress_log["Buffer_Consumption_Percentage"],
        marker="o",
        linestyle="-",
        label="Project Path",
    )

    # Highlight current point
    ax_fever.plot(
        last_snapshot_data["Work_Done_Percentage"],
        last_snapshot_data["Buffer_Consumption_Percentage"],
        "*",
        color="blue",
        markersize=15,
        label=f"Current ({snapshot_date_str})",
    )

    ax_fever.set_xlabel("Work Done (%)")
    ax_fever.set_ylabel("Buffer Consumption (%)")
    ax_fever.set_xlim(0, 1)
    ax_fever.set_ylim(
        0, max(1.1, df_progress_log["Buffer_Consumption_Percentage"].max() * 1.2)
    )
    ax_fever.legend()
    ax_fever.grid(True)
    plt.tight_layout()
    fever_chart_path = f"{snapshot_date_str}_fever_chart.png"
    fig_fever.savefig(fever_chart_path)
    plt.close(fig_fever)
    console.print(f"[green]Fever Chart saved to {fever_chart_path}[/green]")

    return work_execution_chart_path, fever_chart_path


def _insert_charts_into_excel(excel_path: str, chart_paths: tuple[str, str]):
    """Inserts the generated charts into the Excel file."""
    logging.info(f"Inserting charts into {excel_path}")
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image

        book = load_workbook(excel_path)
        work_exec_path, fever_chart_path = chart_paths

        # --- Insert Work Execution Chart ---
        if "Work_Execution_Chart" not in book.sheetnames:
            ws_we = book.create_sheet("Work_Execution_Chart")
        else:
            ws_we = book["Work_Execution_Chart"]
            # Clear existing images
            ws_we._images = []

        img_we = Image(work_exec_path)
        ws_we.add_image(img_we, "A1")

        # --- Insert Fever Chart ---
        if "Fever_Chart" not in book.sheetnames:
            ws_fever = book.create_sheet("Fever_Chart")
        else:
            ws_fever = book["Fever_Chart"]
            ws_fever._images = []

        img_fever = Image(fever_chart_path)
        ws_fever.add_image(img_fever, "A1")

        book.save(excel_path)
        console.print(
            "[bold green]Successfully inserted charts into the Excel workbook.[/bold green]"
        )

    except Exception as e:
        console.print(
            f"[bold red]An error occurred while inserting charts into Excel: {e}[/bold red]"
        )
        logging.error(f"Failed to insert charts: {e}", exc_info=True)
        raise typer.Exit(code=1)


# ==============================================================================
# Main Command
# ==============================================================================


@app.command()
def main(
    excel_path: Optional[str] = typer.Option(
        None, "--excel-path", help="Path to the Excel workbook (.xlsx)."
    ),
    create_template: bool = typer.Option(
        False,
        "--create-template",
        "-c",
        help="Create a new Excel template.",
    ),
    overwrite: bool = typer.Option(
        False,
        "--overwrite",
        "-o",
        help="Overwrite existing Excel file without prompting.",
    ),
    snapshot_date: Optional[str] = typer.Option(
        None,
        "--snapshot-date",
        help="Specifies the single date (YYYY-MM-DD) for processing.",
    ),
    log_level: str = typer.Option(
        "INFO",
        "--log-level",
        help="Sets the logging verbosity.",
        case_sensitive=False,
    ),
    no_chart_insertion: bool = typer.Option(
        False,
        "--no-chart-insertion",
        help="If present, the script will process data but skip inserting charts.",
    ),
    no_data_update: bool = typer.Option(
        False,
        "--no-data-update",
        help="If present, the script will generate charts but skip updating the Progress_Log sheet.",
    ),
):
    """
    Automates the generation of MOVE (Minimal Outcome-Value Effort) project progress reports.
    """
    _setup_logging(log_level)

    if excel_path:
        excel_path = excel_path.strip('"')
    if snapshot_date:
        snapshot_date = snapshot_date.strip('"')

    if create_template:
        logging.info("Template creation requested.")
        output_path = excel_path if excel_path else "MOVE_Project_Template.xlsx"
        _create_excel_template(output_path, overwrite)
        raise typer.Exit()

    # Validate required arguments for processing
    if not excel_path or not snapshot_date:
        console.print(
            "[bold red]Error: --excel-path and --snapshot-date are required unless --create-template is used.[/bold red]"
        )
        raise typer.Exit(code=1)

    logging.info(f"Starting MOVE report generation for: {excel_path}")
    logging.info(f"Snapshot date: {snapshot_date}")

    # 1. Read data
    move_config, df_historic, df_current = _read_excel_data(excel_path)

    # 2. Calculate derived config
    historic_50th_percentile_flow_time = _calculate_historic_flow_time(
        move_config, df_historic
    )

    # 3. Generate progress log
    df_progress_log = _generate_full_progress_log(
        move_config, df_current, historic_50th_percentile_flow_time, snapshot_date
    )

    # 4. Update excel with progress log
    if not no_data_update:
        _update_progress_log_sheet(excel_path, df_progress_log)

    # 5. Generate charts
    chart_paths = None
    if not no_chart_insertion:
        chart_paths = _generate_charts(df_progress_log, move_config, snapshot_date)

    # 6. Insert charts into excel
    if chart_paths:
        _insert_charts_into_excel(excel_path, chart_paths)

    # 7. Cleanup
    if chart_paths:
        import os

        for path in chart_paths:
            try:
                os.remove(path)
                logging.info(f"Cleaned up temporary file: {path}")
            except OSError as e:
                logging.warning(f"Could not remove temporary file {path}: {e}")


def _calculate_historic_flow_time(
    move_config: MOVEConfiguration, df_historic: pd.DataFrame
) -> float:
    """Determines the historic 50th percentile flow time."""
    if move_config.historic_50th_percentile_flow_time_override is not None:
        logging.info(
            f"Using override for historic 50th percentile flow time: {move_config.historic_50th_percentile_flow_time_override}"
        )
        return move_config.historic_50th_percentile_flow_time_override

    if (
        df_historic.empty
        or "Flow_Time_Days" not in df_historic.columns
        or df_historic["Flow_Time_Days"].isnull().all()
    ):
        logging.warning(
            "Historic work items data is empty or lacks 'Flow_Time_Days'. Defaulting to 5 days."
        )
        return 5.0

    # Ensure calculation is robust against NaNs
    flow_time = df_historic["Flow_Time_Days"].dropna()
    if flow_time.empty:
        logging.warning("No valid 'Flow_Time_Days' data found. Defaulting to 5 days.")
        return 5.0

    calculated_percentile = flow_time.quantile(0.5)
    logging.info(
        f"Calculated historic 50th percentile flow time: {calculated_percentile:.2f} days"
    )
    return calculated_percentile


if __name__ == "__main__":
    app()
