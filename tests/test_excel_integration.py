import pytest
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import date, datetime, timedelta
import os
import sys
import subprocess
import random
from dateutil.relativedelta import relativedelta
from matplotlib import pyplot as plt

# Path to the main script (assuming it's in the parent directory)
SCRIPT_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "move_tracker_report.py")
)


@pytest.fixture
def create_test_excel_input(tmp_path):
    """
    Fixture to create a temporary Excel input file for testing,
    mimicking the output of _create_excel_template.
    """
    excel_file = tmp_path / "test_input.xlsx"
    wb = Workbook()

    # --- Fixed dates for reproducible tests ---
    planned_start_date = date(2025, 1, 1)
    planned_delivery_date = date(2025, 2, 1)

    # --- Calculate derived config values in Python ---
    # Historic_50th_Percentile_Flow_Time (from fixed historic data: 4, 7, 2 -> median is 4)
    historic_50th_percentile_flow_time = 4.0

    # Initial_Scope (WI-001 starts on planned_start_date, others later)
    initial_scope = 4

    # Initial_Ideal_Completion_Flow_Time
    initial_ideal_completion_flow_time = (
        initial_scope * historic_50th_percentile_flow_time
    )

    # Buffer Dates (assuming rounding to nearest day for timedelta)
    green_buffer_date = planned_start_date + timedelta(
        days=round(initial_ideal_completion_flow_time)
    )
    yellow_buffer_date = green_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )
    red_buffer_date = yellow_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )
    beyond_red_date = red_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )

    # --- MOVE_Configuration Sheet ---
    ws_config = wb.create_sheet("MOVE_Configuration")
    ws_config.append(["Parameter", "Value"])

    config_data = [
        ("Planned_Start_Date", planned_start_date),
        ("Planned_Delivery_Date", planned_delivery_date),
        ("Historic_50th_Percentile_Flow_Time", historic_50th_percentile_flow_time),
        ("Historic_50th_Percentile_Flow_Time_Override", ""),
        ("Initial_Scope", initial_scope),
        ("Initial_Ideal_Completion_Flow_Time", initial_ideal_completion_flow_time),
        ("Buffer_Green_Date", green_buffer_date),
        ("Buffer_Yellow_Date", yellow_buffer_date),
        ("Buffer_Red_Date", red_buffer_date),
        ("Buffer_Beyond_Red_Date", beyond_red_date),
        ("Fever_Green_Yellow_Left_Y", 0.2),
        ("Fever_Green_Yellow_Right_Y", 0.5),
        ("Fever_Yellow_Red_Left_Y", 0.5),
        ("Fever_Yellow_Red_Right_Y", 0.8),
    ]

    for idx, (param, value) in enumerate(config_data):
        row_num = idx + 2
        ws_config.cell(row=row_num, column=1, value=param)
        ws_config.cell(row=row_num, column=2, value=value)

    # Apply date formatting to buffer dates (B2, B3, B8, B9, B10, B11 based on current config_data)
    ws_config["B2"].number_format = "YYYY-MM-DD"  # Planned_Start_Date
    ws_config["B3"].number_format = "YYYY-MM-DD"  # Planned_Delivery_Date
    ws_config["B8"].number_format = "YYYY-MM-DD"  # Green_Buffer_Date
    ws_config["B9"].number_format = "YYYY-MM-DD"  # Yellow_Buffer_Date
    ws_config["B10"].number_format = "YYYY-MM-DD"  # Red_Buffer_Date
    ws_config["B11"].number_format = "YYYY-MM-DD"  # Beyond_Red_Date

    # --- Historic_Work_Items Sheet ---
    ws_historic = wb.create_sheet("Historic_Work_Items")
    historic_headers = [
        "Historical_WI_ID",
        "Description",
        "Actual_Start_Date",
        "Actual_Completion_Date",
        "Flow_Time_Days",
    ]
    ws_historic.append(historic_headers)

    # Fixed historic data for reproducible tests
    ws_historic.append(
        ["HIST-001", "Sample 1", date(2024, 1, 1), date(2024, 1, 5), 4]
    )  # 4 days
    ws_historic.append(
        ["HIST-002", "Sample 2", date(2024, 1, 10), date(2024, 1, 17), 7]
    )  # 7 days
    ws_historic.append(
        ["HIST-003", "Sample 3", date(2024, 2, 1), date(2024, 2, 3), 2]
    )  # 2 days

    # --- Current_Work_Items Sheet ---
    ws_current = wb.create_sheet("Current_Work_Items")
    current_headers = [
        "Work_Item_ID",
        "Description",
        "Commitment_Date",
        "Status",
        "Actual_Start_Date",
        "Actual_Completion_Date",
        "Date_Withdrawn",
    ]
    ws_current.append(current_headers)

    # Fixed current data for reproducible tests
    ws_current.append(
        [
            "WI-001",
            "Current Item 1",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 1),
            date(2025, 1, 5),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-002",
            "Current Item 2",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 6),
            date(2025, 1, 10),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-003",
            "Current Item 3",
            date(2025, 1, 1),
            "In Progress",
            date(2025, 1, 11),
            None,
            date(2025, 1, 13),
        ]
    )
    ws_current.append(
        [
            "WI-004",
            "Current Item 4",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 14),
            date(2025, 1, 18),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-005",
            "Current Item 5",
            date(2025, 1, 12),
            "Completed",
            date(2025, 1, 17),
            date(2025, 1, 21),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-006",
            "Current Item 6",
            date(2025, 1, 12),
            "Not Started",
            None,
            None,
            date(2025, 1, 18),
        ]
    )

    # --- Other Sheets (empty for now) ---
    wb.create_sheet("Instructions")
    wb.create_sheet("Progress_Log")
    wb.create_sheet("Work_Execution_Chart")
    wb.create_sheet("Fever_Chart")

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(excel_file)
    return excel_file, planned_start_date


def test_progress_log_generation_basic(create_test_excel_input, tmp_path):
    """
    Tests if the Progress_Log sheet is correctly generated for a basic scenario.
    """
    input_excel_path, planned_start_date_fixture = create_test_excel_input
    # snapshot_date = planned_start_date_fixture + timedelta(
    #     days=15
    # )  # Example snapshot date
    snapshot_date = date(2025, 1, 21)
    snapshot_date_str = snapshot_date.strftime("%Y-%m-%d")

    command = [
        sys.executable,
        SCRIPT_PATH,
        "--excel-path",
        str(input_excel_path),
        "--snapshot-date",
        snapshot_date_str,
        "--log-level",
        "DEBUG",
        "--overwrite",
    ]

    result = subprocess.run(command, cwd=tmp_path, capture_output=True, text=True)

    assert (
        result.returncode == 0
    ), f"Script failed with error:\n{result.stderr}\n{result.stdout}"

    # Read the generated Progress_Log sheet
    df_progress_log = pd.read_excel(input_excel_path, sheet_name="Progress_Log")
    # Convert Snapshot_Date column to datetime.date objects to match expected_df
    df_progress_log["Snapshot_Date"] = df_progress_log["Snapshot_Date"].dt.date

    # --- EXPECTED PROGRESS_LOG DATA FOR THIS SCENARIO ---
    # Note: The expected data is based on the fixed historic and current data in the test input.
    expected_data = {
        "Snapshot_Date": [
            date(2025, 1, 1),
            date(2025, 1, 5),
            date(2025, 1, 6),
            date(2025, 1, 10),
            date(2025, 1, 11),
            date(2025, 1, 12),
            date(2025, 1, 13),
            date(2025, 1, 14),
            date(2025, 1, 17),
            date(2025, 1, 18),
            date(2025, 1, 21),
        ],
        "Scope_At_Snapshot": [4, 4, 4, 4, 4, 6, 5, 5, 5, 4, 4],
        "Actual_Work_Completed": [0, 1, 1, 2, 2, 2, 2, 2, 2, 3, 4],
        "Elapsed_Time_Days": [1, 5, 6, 10, 11, 12, 13, 14, 17, 18, 21],
        "Actual_Operational_Throughput": [
            0.0,
            0.2,
            0.16666666666666666,
            0.2,
            0.18181818181818182,
            0.16666666666666666,
            0.15384615384615385,
            0.142857142857143,
            0.117647058823529,
            0.166666666666667,
            0.19047619047619,
        ],
        "Current_50th_Percentile_Flow_Time": [
            4.0,
            4.0,
            4.0,
            5.0,
            5.0,
            5.0,
            5.0,
            5.0,
            5.0,
            5.0,
            5.0,
        ],
        "Forecasted_Delivery_Date": [
            None,
            date(2025, 1, 18),
            date(2025, 1, 20),
            date(2025, 1, 20),
            date(2025, 1, 21),
            date(2025, 2, 2),
            date(2025, 1, 29),
            date(2025, 1, 31),
            date(2025, 2, 5),
            date(2025, 1, 27),
            date(2025, 1, 25),
        ],
        "Buffer_Consumption_Percentage": [
            0,
            0.111111111,
            0.333333333,
            0.333333333,
            0.444444444,
            1.777777778,
            1.333333333,
            1.555555556,
            2.111111111,
            1.111111111,
            0.888888889,
        ],
        "Work_Done_Percentage": [
            0,
            0.25,
            0.25,
            0.5,
            0.5,
            0.333333333,
            0.4,
            0.4,
            0.4,
            0.75,
            1,
        ],
        "Fever_chart_signal": [
            "Green",
            "Green",
            "Yellow",
            "Green",
            "Yellow",
            "Beyond Red",
            "Beyond Red",
            "Beyond Red",
            "Beyond Red",
            "Beyond Red",
            "Red",
        ],
    }
    expected_df = pd.DataFrame(expected_data)
    expected_df["Snapshot_Date"] = pd.to_datetime(expected_df["Snapshot_Date"]).dt.date
    # Convert Forecasted_Delivery_Date in df_progress_log to datetime.date, handling NaT
    df_progress_log["Forecasted_Delivery_Date"] = df_progress_log[
        "Forecasted_Delivery_Date"
    ].apply(lambda x: x.date() if pd.notna(x) else pd.NaT)

    expected_df["Forecasted_Delivery_Date"] = expected_df[
        "Forecasted_Delivery_Date"
    ].apply(
        lambda x: x
        if pd.notna(x)
        else pd.NaT  # Ensure expected_df also has NaT for comparison
    )

    pd.testing.assert_frame_equal(df_progress_log, expected_df, check_dtype=False)

    # Verify chart files are created
    work_exec_chart_path = tmp_path / f"{snapshot_date_str}_work_execution_chart.png"
    fever_chart_path = tmp_path / f"{snapshot_date_str}_fever_chart.png"

    # assert work_exec_chart_path.exists()
    # assert work_exec_chart_path.stat().st_size > 0

    # assert fever_chart_path.exists()
    # assert fever_chart_path.stat().st_size > 0

    # Optional: Verify images are embedded in the Excel file
    wb_output = load_workbook(input_excel_path)
    assert "Work_Execution_Chart" in wb_output.sheetnames
    assert "Fever_Chart" in wb_output.sheetnames
    assert len(wb_output["Work_Execution_Chart"]._images) > 0
    assert len(wb_output["Fever_Chart"]._images) > 0
