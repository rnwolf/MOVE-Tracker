import pytest
from openpyxl import Workbook
from datetime import date, timedelta
import os


@pytest.fixture
def create_test_excel_input(tmp_path):
    """
    Fixture to create a temporary Excel input file for testing,
    mimicking the output of _create_excel_template.
    """
    excel_file = tmp_path / "test_input.xlsx"
    wb = Workbook()

    # --- Fixed dates for reproducible tests ---
    planned_start_date = date(2025, 1, 1)
    planned_delivery_date = date(2025, 2, 1)

    # --- Calculate derived config values in Python ---
    # Historic_50th_Percentile_Flow_Time (from fixed historic data: 4, 7, 2 -> median is 4)
    historic_50th_percentile_flow_time = 4.0

    # Initial_Scope (WI-001 starts on planned_start_date, others later)
    initial_scope = 4

    # Initial_Ideal_Completion_Flow_Time
    initial_ideal_completion_flow_time = (
        initial_scope * historic_50th_percentile_flow_time
    )

    # Buffer Dates (assuming rounding to nearest day for timedelta)
    green_buffer_date = planned_start_date + timedelta(
        days=round(initial_ideal_completion_flow_time)
    )
    yellow_buffer_date = green_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )
    red_buffer_date = yellow_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )
    beyond_red_date = red_buffer_date + timedelta(
        days=round(0.2 * initial_ideal_completion_flow_time)
    )

    # --- MOVE_Configuration Sheet ---
    ws_config = wb.create_sheet("MOVE_Configuration")
    ws_config.append(["Parameter", "Value"])

    config_data = [
        ("Planned_Start_Date", planned_start_date),
        ("Planned_Delivery_Date", planned_delivery_date),
        ("Historic_50th_Percentile_Flow_Time", historic_50th_percentile_flow_time),
        ("Historic_50th_Percentile_Flow_Time_Override", ""),
        ("Initial_Scope", initial_scope),
        ("Initial_Ideal_Completion_Flow_Time", initial_ideal_completion_flow_time),
        ("Buffer_Green_Date", green_buffer_date),
        ("Buffer_Yellow_Date", yellow_buffer_date),
        ("Buffer_Red_Date", red_buffer_date),
        ("Buffer_Beyond_Red_Date", beyond_red_date),
        ("Fever_Green_Yellow_Left_Y", 0.2),
        ("Fever_Green_Yellow_Right_Y", 0.5),
        ("Fever_Yellow_Red_Left_Y", 0.5),
        ("Fever_Yellow_Red_Right_Y", 0.8),
    ]

    for idx, (param, value) in enumerate(config_data):
        row_num = idx + 2
        ws_config.cell(row=row_num, column=1, value=param)
        ws_config.cell(row=row_num, column=2, value=value)

    # Apply date formatting to buffer dates (B2, B3, B8, B9, B10, B11 based on current config_data)
    ws_config["B2"].number_format = "YYYY-MM-DD"  # Planned_Start_Date
    ws_config["B3"].number_format = "YYYY-MM-DD"  # Planned_Delivery_Date
    ws_config["B8"].number_format = "YYYY-MM-DD"  # Green_Buffer_Date
    ws_config["B9"].number_format = "YYYY-MM-DD"  # Yellow_Buffer_Date
    ws_config["B10"].number_format = "YYYY-MM-DD"  # Red_Buffer_Date
    ws_config["B11"].number_format = "YYYY-MM-DD"  # Beyond_Red_Date

    # --- Historic_Work_Items Sheet ---
    ws_historic = wb.create_sheet("Historic_Work_Items")
    historic_headers = [
        "Historical_WI_ID",
        "Description",
        "Actual_Start_Date",
        "Actual_Completion_Date",
        "Flow_Time_Days",
    ]
    ws_historic.append(historic_headers)

    # Fixed historic data for reproducible tests
    ws_historic.append(
        ["HIST-001", "Sample 1", date(2024, 1, 1), date(2024, 1, 5), 4]
    )  # 4 days
    ws_historic.append(
        ["HIST-002", "Sample 2", date(2024, 1, 10), date(2024, 1, 17), 7]
    )  # 7 days
    ws_historic.append(
        ["HIST-003", "Sample 3", date(2024, 2, 1), date(2024, 2, 3), 2]
    )  # 2 days

    # --- Current_Work_Items Sheet ---
    ws_current = wb.create_sheet("Current_Work_Items")
    current_headers = [
        "Work_Item_ID",
        "Description",
        "Commitment_Date",
        "Status",
        "Actual_Start_Date",
        "Actual_Completion_Date",
        "Date_Withdrawn",
    ]
    ws_current.append(current_headers)

    # Fixed current data for reproducible tests
    ws_current.append(
        [
            "WI-001",
            "Current Item 1",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 1),
            date(2025, 1, 5),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-002",
            "Current Item 2",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 6),
            date(2025, 1, 10),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-003",
            "Current Item 3",
            date(2025, 1, 1),
            "In Progress",
            date(2025, 1, 11),
            None,
            date(2025, 1, 13),
        ]
    )
    ws_current.append(
        [
            "WI-004",
            "Current Item 4",
            date(2025, 1, 1),
            "Completed",
            date(2025, 1, 14),
            date(2025, 1, 18),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-005",
            "Current Item 5",
            date(2025, 1, 12),
            "Completed",
            date(2025, 1, 17),
            date(2025, 1, 21),
            None,
        ]
    )
    ws_current.append(
        [
            "WI-006",
            "Current Item 6",
            date(2025, 1, 12),
            "Not Started",
            None,
            None,
            date(2025, 1, 18),
        ]
    )

    # --- Other Sheets (empty for now) ---
    wb.create_sheet("Instructions")
    wb.create_sheet("Progress_Log")
    wb.create_sheet("Work_Execution_Chart")
    wb.create_sheet("Fever_Chart")

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(excel_file)
    return excel_file, planned_start_date